from openpyxl import load_workbook

workbook = load_workbook(filename="files/data.xlsx")
workbook.sheetnames

sheet = workbook.active
# print(sheet)
# print(sheet.title)

# print(sheet["A1"])
# print(sheet["A1"].value)
# print(sheet.cell(row=1, column=1).value)

# print(sheet["A"])
# for item in sheet["C"]:
#     print(item.value)


# for row in sheet.iter_rows(min_row=1,
#                            max_row=4,
#                            min_col=1,
#                            max_col=6):
#     print(row)

# for value in sheet.iter_rows(min_row=1,
#                            max_row=4,
#                            min_col=1,
#                            max_col=6,
#                            values_only=True):
#     print(value)


data = {}
data['exchangeRate'] = []
keys = []
for value in sheet.iter_rows(min_row=1,
                           max_row=1,
                           min_col=1,
                           max_col=6,
                           values_only=True):
    # print(value)
    for id, key in enumerate(value):
       keys.append(key)
# print(keys) 


for values in sheet.iter_rows(min_row=2,
                           max_row=4,
                           min_col=1,
                           max_col=6,
                           values_only=True):
    row = {}
    for id, value in enumerate(values):
       row.update({keys[id]:value})
    data['exchangeRate'].append(row)

print(data)