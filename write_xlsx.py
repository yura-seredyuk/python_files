from openpyxl import Workbook


data = {}
data['exchangeRate'] = []
data['exchangeRate'].append({
    'baseCurrency': 'UAH', 
    'currency': 'USD', 
    'saleRateNB': 27.0275, 
    'purchaseRateNB': 27.0275, 
    'saleRate': 27.18, 
    'purchaseRate': 26.75
})
data['exchangeRate'].append({
    'baseCurrency': 'UAH', 
    'currency': 'PLZ', 
    'saleRateNB': 7.2526, 
    'purchaseRateNB': 7.2526, 
    'saleRate': 7.35, 
    'purchaseRate': 7.05
})
data['exchangeRate'].append({
    'baseCurrency': 'UAH', 
    'currency': 'EUR', 
    'saleRateNB': 32.7722, 
    'purchaseRateNB': 32.7722, 
    'saleRate': 32.95, 
    'purchaseRate': 32.35
})

workbook = Workbook()

sheet = workbook.active
sheet.title = 'PB_data'
for index, item in enumerate(list(data['exchangeRate'][0].keys())):
    sheet.cell(row=1, column=index+1).value = item

row_num = 1
for row in data['exchangeRate']:
    row_num += 1
    for index, key in enumerate(list(row.keys())):
        sheet.cell(row=row_num, column=index+1).value = row[key]
# sheet["A1"] = 'Data1'
# sheet["A2"] = 'Data2'
# sheet.cell(row=1, column=2).value = 'Data_3'



workbook.save(filename="files/data.xlsx")



